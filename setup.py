import json 
import pandas as pd
import openpyxl

db = json.load(open("db.json"))
days_dict = {
    "Monday": {},
    "Tuesday": {},
    "Wednesday": {},
    "Thursday": {},
    "Friday": {},
    "Saturday": {},
}


class Course:
    def __init__(self, code, description="") -> None:
        self.code = code # BITS F111
        self.sections = db["courses"][code]["sections"] if code in db["courses"] else {} # dictionary of section_name: Section.__dict__
        # coomp_code = int 
        # textbooks = []
        # instructors_in_charge = str
        # instructors = []
        # credits = int
        self.description = description
        if self.code not in db["courses"]:
            db["courses"][self.code] = self.__dict__

    def __str__(self):
        """prints out basic information about the class"""
        return self.description
        
    def get_all_sections(self):
        """returns a list of all sections in the class"""
        return list(self.sections.keys())

    def populate_sections(self, section, admin=False):
        """should have access only from admins """
        if admin:
            self.sections[section.section_name] = section.__dict__
            db["courses"][self.code]["sections"][section.section_name] = section.__dict__
            return True
        
        return False
        

# 3 types of sections -> tutorials, labs, lectures
class Section:
    def __init__(self, course, section_name, timings={}):
        self.course = course if type(course)==str else course.code # stores the course code string BITS F111
        # room_no = int
        self.timings = timings # day: hours - {monday: [1, 2, 3] ...}
        self.section_name = section_name
        # course_title = ""
        # instructor = ""
        # students = []
        try:
            parent_course = Course(self.course, db["courses"][self.course]["description"])
            parent_course.populate_sections(self, admin=True)
        except KeyError:
            return f"Course has not been registered yet. Please register {self.course} as a course and then add a section for it."

    def __str__(self):
        return self.section_name

 
class TimeTable():
    def __init__(self, sections=[], tt=days_dict) -> None:
        self.sections = sections # list of Section objects which are added to timetable 
        self.tt = tt    

    def enroll_class(self, section):
        """enroll in a subject"""
        # make sure there are no clashes
        if self.check_clashes(section)!=[]:
            return False

        self.sections.append(section)
        for day in section.timings:
            for hour in section.timings[day]:
                self.tt[day][hour] = section
        return True

    def drop_class(self, section):
        """drop a subject"""
        self.sections.remove(section)
        for day in section.timings:
            for hour in section.timings[day]:
                self.tt[day].pop(hour)

        return True

    def drop_class_by_name(self, section_name):
        # self.courses.pop(i) if section.section_name == section_name else pass for i, section in enumerate(courses)
        section = next(filter(lambda x: x.section_name==section_name, self.sections))
        self.drop_class(section)
        return section

    def check_clashes(self, section): # if this returns [] it means there were no clashes
        clashes = []
        section = section if type(section)==dict else section.timings 
        for day in self.tt:
            # clash= set(self.tt[day]).intersection(set(section[day])) 
            clash = [value for value in self.tt[day] if value in section[day]]
            if clash!=[]:
                clashes.append({day: clash})

        return clashes
                

    def export_to_csv(self):
        """Exports timetable to CSV format and returns timetable DataFrame"""
        data = {}
        data["hour"] = list(range(1, 10))
        for day in list(self.tt.keys()):
            data[day] = [0 for _ in range(1, 10)]
        
        # o(n^3) time complexity skull emoji sorry i haven't slept in 20+ hours
        for section in self.sections:
            for day in list(section.timings.keys()):
                for hour in section.timings[day]:
                    data[day][hour-1] = section.course + " " + section.section_name

        df = pd.DataFrame(data)
        df.to_csv("timetable.csv", header=True, index=False)
        return df

def populate_courses(path_to_sheet, admin=False):
    """adds new courses to the database, takes path to sheet name/name of sheet as input"""
    try:
        courses = openpyxl.load_workbook(path_to_sheet).active
    except:
        return False

    for row in range(1, courses.max_row+1):
        code = courses.cell(row=row, column=1).value
        description = courses.cell(row=row, column=3).value
        Course(code=code, description=description)

    if save_changes():
        return True


def save_changes():
    """saves db to db.json"""
    json.dump(db, open("db.json", "w+"), indent=4)
    return True


# tt = TimeTable()
# course = Course(code="BITS F111", description="Thermodynamics raddi course")
# section = Section(course, "P2", {"Tuesday": [6, 7]})
#
# print(tt.tt)
# print(tt.sections)
#
# tt.drop_class_by_name("P9")

# populate_courses("courses.xlsx", admin=True)

